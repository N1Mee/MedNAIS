#!/usr/bin/env python3
import sys
import json
import os
import asyncio
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Import document processing libraries
import PyPDF2
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PIL import Image
import io

# Import emergentintegrations
from emergentintegrations.llm.chat import LlmChat, UserMessage

def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    try:
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading PDF: {str(e)}"

def extract_text_from_docx(file_path):
    """Extract text from Word document"""
    try:
        doc = DocxDocument(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    except Exception as e:
        return f"Error reading Word document: {str(e)}"

def extract_text_from_excel(file_path):
    """Extract text from Excel file"""
    try:
        wb = load_workbook(file_path, read_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

def extract_text_from_image(file_path):
    """For images, we'll send the file path and let GPT-5 Vision handle it"""
    try:
        # Verify it's a valid image
        img = Image.open(file_path)
        return f"[Image file: {img.format}, Size: {img.size}]"
    except Exception as e:
        return f"Error reading image: {str(e)}"

def extract_content(file_path, mime_type):
    """Extract content based on file type"""
    file_path_lower = file_path.lower()
    
    if mime_type.startswith('image/') or file_path_lower.endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
        return extract_text_from_image(file_path), 'image'
    elif mime_type == 'application/pdf' or file_path_lower.endswith('.pdf'):
        return extract_text_from_pdf(file_path), 'text'
    elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword'] or file_path_lower.endswith(('.docx', '.doc')):
        return extract_text_from_docx(file_path), 'text'
    elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel'] or file_path_lower.endswith(('.xlsx', '.xls')):
        return extract_text_from_excel(file_path), 'text'
    elif mime_type.startswith('text/') or file_path_lower.endswith('.txt'):
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read(), 'text'
    else:
        # Try as text by default
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read(), 'text'
        except:
            return f"Unsupported file type: {mime_type}", 'text'

async def generate_sop_steps(content, content_type, file_path=None, custom_prompt=None):
    """Use GPT-5 to generate SOP steps from content"""
    
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        raise Exception("OPENAI_API_KEY not found in environment variables")
    
    # Use custom prompt instructions if provided
    if custom_prompt:
        system_message = f"""You are an expert at creating Standard Operating Procedures (SOPs). 
Your task is to analyze documents and convert them into clear, actionable SOP steps.

{custom_prompt}"""
    else:
        system_message = """You are an expert at creating Standard Operating Procedures (SOPs). 
Your task is to analyze documents and convert them into clear, actionable SOP steps.

Each step should have:
- A clear, concise title (max 60 characters)
- Detailed description of what to do
- Optional timer in seconds if the step requires specific timing
- Optional references or notes

Return ONLY a valid JSON array of steps in this exact format:
[
  {
    "title": "Step title",
    "description": "Detailed description of what to do in this step",
    "timerSeconds": 300,
    "references": ["Reference 1", "Reference 2"]
  }
]

Important:
- Return ONLY the JSON array, no additional text
- Create 3-15 steps depending on complexity
- Be specific and actionable
- Include timing when relevant
- Make it easy to follow"""
    
    # Initialize LLM Chat with GPT-5
    chat = LlmChat(
        api_key=api_key,
        session_id=f"sop-generation-{os.getpid()}",
        system_message=system_message
    ).with_model("openai", "gpt-5")
    
    # Create prompt based on content type
    if content_type == 'image' and file_path:
        prompt = f"""Analyze this image and create SOP steps based on what you see.
        
The image contains: {content}

Please create detailed SOP steps that describe the process shown in the image.
Return ONLY the JSON array of steps."""
    else:
        prompt = f"""Analyze this document and create SOP steps:

{content[:8000]}  

Please create detailed SOP steps based on this content.
Return ONLY the JSON array of steps."""
    
    # Send message and get response
    user_message = UserMessage(text=prompt)
    response = await chat.send_message(user_message)
    
    return response

async def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "Usage: python process_document.py <file_path> <mime_type> [custom_prompt_file]"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    mime_type = sys.argv[2]
    custom_prompt_file = sys.argv[3] if len(sys.argv) > 3 else None
    
    try:
        # Extract content
        content, content_type = extract_content(file_path, mime_type)
        
        if not content or content.startswith("Error"):
            raise Exception(content)
        
        # Read custom prompt if provided
        custom_prompt = None
        if custom_prompt_file and os.path.exists(custom_prompt_file):
            with open(custom_prompt_file, 'r', encoding='utf-8') as f:
                custom_prompt = f.read()
        
        # Generate SOP steps using GPT-5
        response = await generate_sop_steps(
            content, 
            content_type, 
            file_path if content_type == 'image' else None,
            custom_prompt
        )
        
        # Try to parse as JSON
        try:
            # Clean response - remove markdown code blocks if present
            cleaned_response = response.strip()
            if cleaned_response.startswith('```'):
                # Remove markdown code block markers
                lines = cleaned_response.split('\n')
                cleaned_response = '\n'.join(lines[1:-1] if len(lines) > 2 else lines)
            
            steps = json.loads(cleaned_response)
            
            # Validate structure
            if not isinstance(steps, list):
                raise Exception("Response is not a JSON array")
            
            # Add IDs and order
            for i, step in enumerate(steps):
                step['id'] = f"step-{i+1}"
                step['order'] = i + 1
                
            print(json.dumps(steps))
            
        except json.JSONDecodeError as e:
            # If JSON parsing fails, try to extract JSON from response
            import re
            json_match = re.search(r'\[.*\]', response, re.DOTALL)
            if json_match:
                steps = json.loads(json_match.group())
                for i, step in enumerate(steps):
                    step['id'] = f"step-{i+1}"
                    step['order'] = i + 1
                print(json.dumps(steps))
            else:
                raise Exception(f"Could not parse JSON from response: {response[:500]}")
        
    except Exception as e:
        error_response = {
            "error": str(e),
            "steps": [
                {
                    "id": "step-1",
                    "order": 1,
                    "title": "Error processing document",
                    "description": f"Could not process the document: {str(e)}",
                    "timerSeconds": 0
                }
            ]
        }
        print(json.dumps(error_response.get("steps")))
        sys.exit(0)

if __name__ == "__main__":
    asyncio.run(main())
